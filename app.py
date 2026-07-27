"""
Freight Billing Report — a Streamlit dashboard.

Joins a billing export (one row per charge line) to a shipment listing report
(one row per shipment), then renders on-screen summaries and produces a
downloadable multi-sheet Excel report.

Everything deployment-specific — app title, theme colours, currency codes and
the source-system column names — lives in the CONFIG block below. Nothing in
the logic assumes a particular organisation, so this file can be published as
is and adapted by changing the constants.
"""

import hashlib
import hmac
import io
from string import Template

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Font, PatternFill

# ── CONFIG ────────────────────────────────────────────────────────────────────

APP_TITLE = "Freight Billing Report"
APP_SUBTITLE = "Shipment & Charge Consolidation"
REPORT_FILENAME = "billing_report.xlsx"

# Currency the report is totalled in, and the secondary currency reported
# alongside it. Amounts in any currency still roll up into the base via the
# billing export's own local-total column.
BASE_CCY = "AUD"
FX_CCY = "USD"
LOCAL_LABEL = f"Local Total ({BASE_CCY})"

# Names and formatting that appear in the exported workbook. These are kept
# byte-identical to the established report format so downstream spreadsheets
# and any formulas pointing at them keep working.
SUPPLIER_TOTAL_COL = f"Local_Total_{BASE_CCY}"
REPORT_HEADER_FILL = "1A56A0"  # Analysis section header fill

# Column names as they appear in the two source exports.
BILLING_JOB_COL = "Job"          # job/shipment key in the billing export
SHIPMENT_KEY_COL = "Shipment"    # job/shipment key in the shipment listing
SHIPMENT_USECOLS = range(1, 22)  # columns to read from the shipment listing
JOB_ID_PATTERN = r"^[A-Za-z]{1,3}\d+"  # what a job reference looks like

# Source column -> internal name. Add or edit entries to match your export.
SHIPMENT_COLUMN_MAP = {
    SHIPMENT_KEY_COL: "Shipment Job",
    "Order Ref": "Order Reference",
    "INCO": "Incoterms",
    "Pack Mode": "Mode",
    "Consignor Name": "Supplier Name",
    "Origin": "Loading Port",
    "Dest.": "Destination Port",
    "Carrier Booking Reference": "Booking Ref",
    "Vessel": "Vessel / Voyage",
    "No. of Cont.": "Container Count",
    "Container #": "Containers",
}
DATE_COLS = ("ETD", "ETA", "ATD", "ATA")

# ── Invoice summary (charge-by-container) config ──────────────────────────────
# Billing-export columns the allocation reads.
CHARGE_CODE_COL = "Charges"        # short charge code, e.g. DCART, OTHC
CHARGE_AMOUNT_COL = "Local Amount"  # GST-exclusive amount in the base currency
CONTAINER_LIST_COL = "Containers"   # "ABCD1234567 (40HC), EFGH7654321 (40HC)"

# The export may or may not carry invoice identifiers. First match wins;
# comparison is case-insensitive and matches on substring.
INVOICE_NUMBER_CANDIDATES = ("invoice number", "invoice no", "invoice #",
                             "transaction number", "commercial invoice")
INVOICE_DATE_CANDIDATES = ("invoice date", "transaction date", "tax date")

# Output headers. Change these if a downstream template expects other names.
JOB_LABEL = "Booking Ref"
CONTAINER_LABEL = "Container"
COMMERCIAL_REF_LABEL = "SBFO Ref (Commercial Invoice)"
INVOICE_NUMBER_LABEL = "Invoice Number"
INVOICE_DATE_LABEL = "Invoice Date"
NO_CONTAINER_LABEL = "(no container listed)"
INVOICE_CSV_FILENAME = "charge_by_container.csv"

# Theme — neutral defaults, no brand assets or brand colours.
PRIMARY = "#33475B"
PRIMARY_LIGHT = "#48627E"
ACCENT = "#0E8074"
ACCENT_DARK = "#0A6158"
BG = "#EDF1F5"
SURFACE = "#FFFFFF"
TEXT = "#1B2A38"
MUTED = "#7A8CA0"
BORDER = "#D3DCE6"
GRID = "#E6EDF3"
CONTROL_BG = "#3E5670"

PALETTE = [ACCENT, PRIMARY, "#5B8FB9", "#8FBF9F", "#B08EA2", "#C98B5E", "#6E7B8B"]

st.set_page_config(
    page_title=APP_TITLE,
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Styling ───────────────────────────────────────────────────────────────────

CSS = Template("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background-color: $bg; color: $text; }
[data-testid="stSidebar"] { background-color: $primary !important; border-right: none; }
[data-testid="stSidebar"] * { color: #ffffff !important; }
[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] { background-color: $accent !important; color: #ffffff !important; }
[data-testid="stSidebar"] [data-baseweb="select"] { background-color: $control !important; }
[data-testid="stSidebar"] [data-baseweb="select"] * { background-color: $control !important; color: #ffffff !important; border-color: $primary_light !important; }
.app-header { background: linear-gradient(90deg,$primary 0%,$primary_light 100%); border-bottom:4px solid $accent; padding:20px 32px; margin:-1rem -1rem 2rem -1rem; display:flex; align-items:center; gap:20px; }
.app-header h1 { font-size:1.7rem; font-weight:700; color:#fff; margin:0; letter-spacing:.01em; }
.app-subtitle { font-size:.78rem; color:#c8d6e5; letter-spacing:.08em; text-transform:uppercase; margin-top:4px; }
.section-title { font-size:.95rem; font-weight:600; color:$primary; text-transform:uppercase; letter-spacing:.08em; border-left:4px solid $accent; padding-left:10px; margin:20px 0 10px 0; }
[data-testid="stDataFrame"] { border:1px solid $border; border-radius:6px; background-color:$surface; }
.stTabs [data-baseweb="tab-list"] { background-color:$surface; border-bottom:2px solid $border; gap:0; }
.stTabs [data-baseweb="tab"] { font-weight:600; font-size:.85rem; letter-spacing:.04em; text-transform:uppercase; color:$muted !important; background:transparent !important; border:none !important; padding:12px 24px; }
.stTabs [aria-selected="true"] { color:$primary !important; border-bottom:3px solid $accent !important; }
.stDownloadButton button, .stButton button { background-color:$accent !important; color:#fff !important; font-weight:600 !important; letter-spacing:.06em !important; text-transform:uppercase !important; border:none !important; border-radius:4px !important; padding:10px 28px !important; }
.stDownloadButton button:hover, .stButton button:hover { background-color:$accent_dark !important; }
[data-testid="stFileUploader"] { background-color:$control; border:2px dashed $primary_light; border-radius:6px; padding:8px; }
[data-testid="metric-container"] { background-color:$surface; border:1px solid $border; border-top:4px solid $accent; padding:16px; border-radius:6px; box-shadow:0 2px 8px rgba(27,42,56,.08); }
[data-testid="stMetricValue"] { font-size:1.8rem !important; font-weight:700 !important; color:$primary !important; }
[data-testid="stMetricLabel"] { color:$muted !important; font-size:.7rem !important; text-transform:uppercase; letter-spacing:.1em; }
div[data-testid="stVerticalBlock"] { gap:0.5rem; }
.stAlert { background-color:$grid !important; border-color:$primary !important; color:$text !important; }
[data-testid="stToolbar"] { display: none !important; }
header[data-testid="stHeader"] { background: transparent !important; }
/* Keep the sidebar visible and expanded at all times */
[data-testid="stSidebar"],
[data-testid="stSidebar"][aria-expanded="false"] {
    transform: none !important;
    visibility: visible !important;
    margin-left: 0 !important;
    min-width: 244px !important;
    width: 244px !important;
}
[data-testid="stSidebarCollapseButton"] { display: none !important; }
</style>
""").safe_substitute(
    bg=BG, text=TEXT, primary=PRIMARY, primary_light=PRIMARY_LIGHT,
    accent=ACCENT, accent_dark=ACCENT_DARK, surface=SURFACE, muted=MUTED,
    border=BORDER, grid=GRID, control=CONTROL_BG,
)
st.markdown(CSS, unsafe_allow_html=True)

PLOTLY_LAYOUT = dict(
    paper_bgcolor=SURFACE, plot_bgcolor="#F7FAFC",
    font=dict(family="Inter, sans-serif", color=TEXT, size=11),
    xaxis=dict(gridcolor=GRID, linecolor=BORDER, tickcolor=TEXT, tickfont=dict(color=TEXT)),
    yaxis=dict(gridcolor=GRID, linecolor=BORDER, tickcolor=TEXT, tickfont=dict(color=TEXT)),
    margin=dict(l=20, r=20, t=36, b=20),
    legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=TEXT)),
)


# ── Access control ────────────────────────────────────────────────────────────
#
# Credentials are NEVER stored in this file. They live in .streamlit/secrets.toml
# locally (git-ignored) or in the host's secrets manager when deployed, as
# salted PBKDF2-SHA256 hashes:
#
#   [users]
#   someuser = "pbkdf2_sha256$240000$<salt hex>$<hash hex>"
#
# Generate a hash with tools/hash_password.py. If no users are configured the
# app refuses all access rather than falling open.

MAX_ATTEMPTS = 6


def load_users():
    try:
        return dict(st.secrets["users"])
    except Exception:
        return {}


def password_matches(stored, supplied):
    """Constant-time check of a supplied password against a stored hash."""
    try:
        scheme, iterations, salt_hex, digest_hex = str(stored).split("$")
        if scheme != "pbkdf2_sha256":
            return False
        calculated = hashlib.pbkdf2_hmac(
            "sha256", supplied.encode("utf-8"), bytes.fromhex(salt_hex), int(iterations)
        )
    except (ValueError, TypeError):
        return False
    return hmac.compare_digest(calculated.hex(), digest_hex)


def login_screen():
    """Render the sign-in form. Never returns if the visitor is not signed in."""
    st.markdown(
        '<style>[data-testid="stSidebar"], [data-testid="stSidebar"][aria-expanded="false"]'
        '{ display:none !important; }</style>',
        unsafe_allow_html=True,
    )
    st.markdown(f"""
    <div class="app-header">
      <div>
        <h1>{APP_TITLE}</h1>
        <div class="app-subtitle">Sign in to continue</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    users = load_users()
    _, mid, _ = st.columns([1, 1.4, 1])
    with mid:
        if not users:
            st.error("No accounts are configured. Set the [users] section in the app's secrets.")
            st.stop()

        if st.session_state.get("login_attempts", 0) >= MAX_ATTEMPTS:
            st.error("Too many failed attempts. Reload the page to try again.")
            st.stop()

        with st.form("sign_in"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Sign in")

        if submitted:
            stored = users.get(username.strip())
            if stored is not None and password_matches(stored, password):
                st.session_state["auth_user"] = username.strip()
                st.session_state["login_attempts"] = 0
                st.rerun()
            st.session_state["login_attempts"] = st.session_state.get("login_attempts", 0) + 1
            # Deliberately vague: don't reveal which of the two was wrong.
            st.error("Incorrect username or password.")

        st.caption("Access is restricted. Contact the report owner if you need an account.")
    st.stop()


def require_login():
    if not st.session_state.get("auth_user"):
        login_screen()


require_login()


# ── Loading ───────────────────────────────────────────────────────────────────

def read_billing(file_obj):
    try:
        return pd.read_excel(file_obj)
    except Exception as e:
        st.error(f"Could not read the billing export: {e}")
        return None


def read_shipment(file_obj):
    """Read the shipment listing, locating the header row before parsing."""
    try:
        raw = pd.read_excel(file_obj, header=None)
        header_row = 0
        for i, row in raw.iterrows():
            vals = [str(v).strip() for v in row if str(v).strip() not in ("", "nan")]
            if SHIPMENT_KEY_COL in vals:
                header_row = i
                break
        file_obj.seek(0)
        df = pd.read_excel(file_obj, header=header_row, usecols=SHIPMENT_USECOLS)
        df.columns = [str(c).strip() for c in df.columns]
        if SHIPMENT_KEY_COL in df.columns:
            df = df[df[SHIPMENT_KEY_COL].astype(str).str.match(JOB_ID_PATTERN)]
            df = df.reset_index(drop=True)
            df = df.rename(columns={k: v for k, v in SHIPMENT_COLUMN_MAP.items() if k in df.columns})
            for col in DATE_COLS:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
            if "Container Count" in df.columns:
                df["Container Count"] = (
                    pd.to_numeric(df["Container Count"], errors="coerce").fillna(0).astype(int)
                )
        return df
    except Exception as e:
        st.error(f"Could not read the shipment listing: {e}")
        return None


def load_data(billing_files, shipment_file):
    sheets = [read_billing(f) for f in billing_files]
    sheets = [s for s in sheets if s is not None]
    if not sheets:
        return None
    billing = pd.concat(sheets, ignore_index=True)
    shipment = read_shipment(shipment_file)
    if shipment is None:
        return None
    merged = pd.merge(
        billing, shipment, how="left",
        left_on=BILLING_JOB_COL, right_on="Shipment Job",
    )
    return merged, shipment


# ── Table builders ────────────────────────────────────────────────────────────

def build_shipment_summary(shipment_df):
    cols = ["Shipment Job", "Supplier Name", "Loading Port", "Destination Port",
            "Order Reference", "Incoterms", "Containers", "Container Count",
            "Mode", "Vessel / Voyage", "Booking Ref", *DATE_COLS]
    return shipment_df[[c for c in cols if c in shipment_df.columns]].copy().reset_index(drop=True)


def build_billing_detail(data):
    cols = ["Shipment Job", "Order Reference", "Supplier Name", "Loading Port",
            "Destination Port", "Incoterms", "Containers", "Description", "Currency",
            "Amount", "Tax", "Total", "Local Total", "Exchange Rate"]
    return data[[c for c in cols if c in data.columns]].copy().reset_index(drop=True)


def build_billing_summary(data, shipment_df):
    base = (data[data["Currency"] == BASE_CCY].groupby("Shipment Job")["Total"].sum()
            .reset_index().rename(columns={"Total": BASE_CCY}))
    fx = (data[data["Currency"] == FX_CCY].groupby("Shipment Job")["Total"].sum()
          .reset_index().rename(columns={"Total": FX_CCY}))
    local = (data.groupby("Shipment Job")["Local Total"].sum()
             .reset_index().rename(columns={"Local Total": LOCAL_LABEL}))
    cur = pd.merge(base, fx, how="outer", on="Shipment Job").fillna(0)
    cur = pd.merge(cur, local, how="left", on="Shipment Job").fillna(0)
    ctx_cols = ["Shipment Job", "Supplier Name", "Loading Port", "Destination Port",
                "Order Reference", "Incoterms", "Containers"]
    ctx = shipment_df[[c for c in ctx_cols if c in shipment_df.columns]].drop_duplicates("Shipment Job")
    return pd.merge(ctx, cur, how="right", on="Shipment Job").reset_index(drop=True)


def build_supplier_summary(data, shipment_df):
    bd = build_billing_detail(data)
    sup = bd.groupby("Supplier Name").agg(
        Shipments=("Shipment Job", "nunique"),
        Charge_Lines=("Shipment Job", "count"),
        **{SUPPLIER_TOTAL_COL: ("Local Total", "sum")},
    ).reset_index()
    if "Container Count" in shipment_df.columns:
        cnt = shipment_df.groupby("Supplier Name")["Container Count"].sum().reset_index()
        sup = pd.merge(sup, cnt, how="left", on="Supplier Name")
    return sup.sort_values(SUPPLIER_TOTAL_COL, ascending=False).reset_index(drop=True)


def add_month_col(df, ship_sum):
    """Add a Month label and sort key to a billing detail frame, keyed off ETD."""
    if "ETD" not in ship_sum.columns:
        return df, []
    etd_map = ship_sum.set_index("Shipment Job")["ETD"]
    out = df.copy()
    out["_etd"] = pd.to_datetime(out["Shipment Job"].map(etd_map), errors="coerce")
    out["_sort"] = out["_etd"].dt.to_period("M").dt.to_timestamp()
    out["Month"] = out["_etd"].dt.strftime("%b %Y")
    order = out[["Month", "_sort"]].drop_duplicates().sort_values("_sort")["Month"].tolist()
    return out, order


# ── Invoice summary: charge by container ──────────────────────────────────────

def find_column(df, candidates):
    """First column whose name contains one of the candidate substrings."""
    for candidate in candidates:
        for col in df.columns:
            if candidate in str(col).lower():
                return col
    return None


def parse_containers(value):
    """'ABCD1234567 (40HC), EFGH7654321 (40HC)' -> ['ABCD1234567', 'EFGH7654321']."""
    if pd.isna(value):
        return []
    out = []
    for part in str(value).replace(";", ",").split(","):
        number = part.split("(")[0].strip()
        if number and number.lower() != "nan":
            out.append(number)
    return out


def container_map(shipment_df):
    """Shipment job -> ordered list of container numbers."""
    if CONTAINER_LIST_COL not in shipment_df.columns:
        return {}
    mapping = {}
    for job, value in zip(shipment_df["Shipment Job"], shipment_df[CONTAINER_LIST_COL]):
        containers = parse_containers(value)
        if containers:
            # Preserve order, drop repeats if a job appears on several rows.
            existing = mapping.setdefault(job, [])
            existing.extend(c for c in containers if c not in existing)
    return mapping


def split_amount(total, parts):
    """Split an amount into `parts` shares that sum back exactly to the total.

    Works in whole cents and hands the leftover cents to the earliest shares
    (largest-remainder), so 218.00 over 11 containers becomes nine at 19.82
    and two at 19.81 rather than eleven at 19.82.
    """
    if parts <= 0:
        return []
    cents = int(round(float(total) * 100))
    sign = -1 if cents < 0 else 1
    base, remainder = divmod(abs(cents), parts)
    return [sign * (base + (1 if i < remainder else 0)) / 100 for i in range(parts)]


def build_charge_by_container(billing_df, shipment_df):
    """One row per invoice + container, one column per charge code.

    Every charge line is allocated evenly across the containers on its job.
    Returns (wide dataframe, diagnostics dict).
    """
    diagnostics = {"missing_columns": [], "jobs_without_containers": [],
                   "invoice_number_source": None, "invoice_date_source": None}

    for col in (CHARGE_CODE_COL, CHARGE_AMOUNT_COL, BILLING_JOB_COL):
        if col not in billing_df.columns:
            diagnostics["missing_columns"].append(col)
    if diagnostics["missing_columns"]:
        return pd.DataFrame(), diagnostics

    invoice_no_col = find_column(billing_df, INVOICE_NUMBER_CANDIDATES)
    invoice_date_col = find_column(billing_df, INVOICE_DATE_CANDIDATES)
    diagnostics["invoice_number_source"] = invoice_no_col
    diagnostics["invoice_date_source"] = invoice_date_col

    containers_by_job = container_map(shipment_df)

    work = billing_df.copy()
    work["_job"] = work[BILLING_JOB_COL].astype(str).str.strip()
    work["_code"] = work[CHARGE_CODE_COL].astype(str).str.strip().str.upper()
    work["_amount"] = pd.to_numeric(work[CHARGE_AMOUNT_COL], errors="coerce").fillna(0.0)
    work["_invoice"] = (work[invoice_no_col].astype(str).str.strip()
                        if invoice_no_col else "")
    if invoice_date_col:
        work["_invoice_date"] = pd.to_datetime(work[invoice_date_col], errors="coerce")
    else:
        work["_invoice_date"] = pd.NaT
    work = work[work["_amount"] != 0]

    # One figure per invoice + job + charge code before allocating.
    grouped = (work.groupby(["_invoice", "_job", "_code"], dropna=False)
               .agg(amount=("_amount", "sum"), invoice_date=("_invoice_date", "max"))
               .reset_index()
               .rename(columns={"_invoice": "invoice", "_job": "job", "_code": "code"}))

    records = {}
    for row in grouped.itertuples(index=False):
        containers = containers_by_job.get(row.job) or containers_by_job.get(row.job.upper())
        if not containers:
            containers = [NO_CONTAINER_LABEL]
            if row.job not in diagnostics["jobs_without_containers"]:
                diagnostics["jobs_without_containers"].append(row.job)
        for container, share in zip(containers, split_amount(row.amount, len(containers))):
            key = (row.invoice, row.job, container, row.invoice_date)
            record = records.setdefault(key, {})
            record[row.code] = round(record.get(row.code, 0.0) + share, 2)

    if not records:
        return pd.DataFrame(), diagnostics

    rows = []
    for (invoice, job, container, invoice_date), charges in records.items():
        row = {JOB_LABEL: job, CONTAINER_LABEL: container,
               COMMERCIAL_REF_LABEL: invoice, INVOICE_NUMBER_LABEL: invoice,
               INVOICE_DATE_LABEL: (invoice_date.strftime("%d/%m/%Y")
                                    if pd.notna(invoice_date) else "")}
        row.update(charges)
        rows.append(row)

    wide = pd.DataFrame(rows)
    code_cols = sorted(c for c in wide.columns if c not in (
        JOB_LABEL, CONTAINER_LABEL, COMMERCIAL_REF_LABEL,
        INVOICE_NUMBER_LABEL, INVOICE_DATE_LABEL))
    wide["TOTAL"] = wide[code_cols].sum(axis=1).round(2)
    wide = wide[[JOB_LABEL, CONTAINER_LABEL, COMMERCIAL_REF_LABEL, *code_cols,
                 "TOTAL", INVOICE_NUMBER_LABEL, INVOICE_DATE_LABEL]]
    return (wide.sort_values([INVOICE_NUMBER_LABEL, JOB_LABEL, CONTAINER_LABEL])
            .reset_index(drop=True), diagnostics)


def reorder_to_template(wide, template_columns):
    """Match a template's column order. Columns the template doesn't mention —
    charge codes new since the template was made — keep their place at the end
    rather than being dropped."""
    tail = [c for c in ("TOTAL", INVOICE_NUMBER_LABEL, INVOICE_DATE_LABEL)
            if c in wide.columns]
    ordered = []
    for col in template_columns:
        if col in wide.columns and col not in ordered and col not in tail:
            ordered.append(col)
    extras = [col for col in wide.columns if col not in ordered and col not in tail]
    return wide.reindex(columns=ordered + extras + tail)


# ── Excel report ──────────────────────────────────────────────────────────────

def build_analysis_summary(billing_det, shipment_sum):
    """Build the flat analysis sections written to the Excel report."""
    sections = {}

    containers = ""
    if "Container Count" in shipment_sum.columns:
        containers = int(shipment_sum["Container Count"].fillna(0).sum())

    sections["kpis"] = pd.DataFrame([
        {"Metric": "Total Shipments", "Value": billing_det["Shipment Job"].nunique()},
        {"Metric": "Total Containers", "Value": containers},
        {"Metric": f"{BASE_CCY} Charges",
         "Value": billing_det[billing_det["Currency"] == BASE_CCY]["Total"].sum()},
        {"Metric": f"{FX_CCY} Charges",
         "Value": billing_det[billing_det["Currency"] == FX_CCY]["Total"].sum()},
        {"Metric": f"Total Billed ({BASE_CCY})", "Value": billing_det["Local Total"].sum()},
    ])

    sup = (billing_det.groupby("Supplier Name")["Local Total"].sum()
           .sort_values(ascending=False).reset_index())
    sup.columns = ["Supplier Name", LOCAL_LABEL]
    sections["supplier"] = sup

    bd_m, _ = add_month_col(billing_det, shipment_sum)
    if bd_m is not None and "Month" in bd_m.columns:
        monthly = (bd_m.groupby(["Month", "Currency"])["Local Total"].sum()
                   .unstack(fill_value=0).reset_index())
        monthly.columns.name = None
        sections["monthly"] = monthly

    top10 = (billing_det.groupby("Description")["Local Total"].sum()
             .sort_values(ascending=False).head(10).reset_index())
    top10.columns = ["Charge Description", LOCAL_LABEL]
    sections["top10"] = top10

    if "Incoterms" in billing_det.columns:
        inco = (billing_det.groupby("Incoterms")["Local Total"].sum()
                .sort_values(ascending=False).reset_index())
        inco.columns = ["Incoterms", LOCAL_LABEL]
        sections["incoterm"] = inco

    return sections


def create_report(shipment_sum, billing_sum, billing_det, supplier_sum):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        billing_det.to_excel(writer, sheet_name="Billing Detail", index=False)
        shipment_sum.to_excel(writer, sheet_name="Shipment Summary", index=False)
        billing_sum.to_excel(writer, sheet_name="Billing Summary", index=False)
        supplier_sum.to_excel(writer, sheet_name="Supplier Summary", index=False)

        analysis = build_analysis_summary(billing_det, shipment_sum)
        ws = writer.book.create_sheet("Analysis")
        row = 1

        section_labels = {
            "kpis": "KEY METRICS",
            "supplier": f"SPEND BY SUPPLIER (Local {BASE_CCY})",
            "monthly": f"MONTHLY CHARGES — {BASE_CCY} vs {FX_CCY} (Local {BASE_CCY})",
            "top10": f"TOP 10 CHARGE TYPES (Local {BASE_CCY})",
            "incoterm": f"SPEND BY INCOTERM (Local {BASE_CCY})",
        }

        for key, label in section_labels.items():
            if key not in analysis:
                continue
            df = analysis[key]
            header_cell = ws.cell(row=row, column=1, value=label)
            header_cell.font = Font(bold=True, color="FFFFFF", size=11)
            header_cell.fill = PatternFill("solid", start_color=REPORT_HEADER_FILL)
            row += 1
            for col_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=row, column=col_idx, value=col_name).font = Font(bold=True)
            row += 1
            for _, data_row in df.iterrows():
                for col_idx, val in enumerate(data_row, start=1):
                    ws.cell(row=row, column=col_idx, value=val)
                row += 1
            row += 1  # blank line between sections

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

    buf.seek(0)
    return buf


# ── Header ────────────────────────────────────────────────────────────────────

st.markdown(f"""
<div class="app-header">
  <div>
    <h1>{APP_TITLE}</h1>
    <div class="app-subtitle">{APP_SUBTITLE}</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Uploads ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown('<div class="section-title" style="color:#fff;border-color:'
                + ACCENT + '">Upload files</div>', unsafe_allow_html=True)
    billing_uploads = st.file_uploader(
        "Billing export(s)", type=["xlsx", "csv"], accept_multiple_files=True)
    shipment_upload = st.file_uploader("Shipment listing report", type=["xlsx", "csv"])
    st.caption(f"Signed in as {st.session_state['auth_user']}")
    if st.button("Sign out"):
        st.session_state.clear()
        st.rerun()

if not billing_uploads or not shipment_upload:
    st.info("Upload the billing export(s) and the shipment listing report in the sidebar to begin.")
    st.stop()

with st.spinner("Processing files…"):
    result = load_data(billing_uploads, shipment_upload)
if result is None:
    st.error("The files could not be loaded. Check that both exports match the expected columns.")
    st.stop()

data, shipment_df = result
ship_sum = build_shipment_summary(shipment_df)
bill_det = build_billing_detail(data)
bill_sum = build_billing_summary(data, shipment_df)
supp_sum = build_supplier_summary(data, shipment_df)
billed_jobs = set(data["Shipment Job"].dropna())

bd_all, all_months = add_month_col(bill_det, ship_sum)

# ── Filters ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("---")
    st.markdown('<div class="section-title" style="color:#fff;border-color:'
                + ACCENT + '">Filters</div>', unsafe_allow_html=True)
    billed_ship = ship_sum[ship_sum["Shipment Job"].isin(billed_jobs)]

    def options(col):
        if col not in billed_ship.columns:
            return []
        return sorted(billed_ship[col].dropna().unique())

    f_month = st.multiselect("Month (ETD)", all_months, placeholder="All months")
    f_supplier = st.multiselect("Supplier", options("Supplier Name"), placeholder="All suppliers")
    f_origin = st.multiselect("Loading port", options("Loading Port"), placeholder="All loading ports")
    f_dest = st.multiselect("Destination port", options("Destination Port"), placeholder="All destinations")
    f_inco = st.multiselect("Incoterms", options("Incoterms"), placeholder="All Incoterms")
    f_mode = st.multiselect("Mode", options("Mode"), placeholder="All modes")
    st.markdown("---")
    st.markdown('<div class="section-title" style="color:#fff;border-color:'
                + ACCENT + '">Download</div>', unsafe_allow_html=True)


def filter_df(df):
    m = pd.Series(True, index=df.index)
    for values, col in (
        (f_supplier, "Supplier Name"),
        (f_origin, "Loading Port"),
        (f_dest, "Destination Port"),
        (f_inco, "Incoterms"),
        (f_mode, "Mode"),
    ):
        if values and col in df.columns:
            m &= df[col].isin(values)
    return df[m]


bill_sum_f = filter_df(bill_sum)

if f_month:
    jobs_in_month = bd_all[bd_all["Month"].isin(f_month)]["Shipment Job"].unique()
    bill_sum_f = bill_sum_f[bill_sum_f["Shipment Job"].isin(jobs_in_month)]

kept = bill_sum_f["Shipment Job"]
ship_sum_f = ship_sum[ship_sum["Shipment Job"].isin(kept)]
bill_det_f = bill_det[bill_det["Shipment Job"].isin(kept)]
fdata = data[data["Shipment Job"].isin(kept)]
fship = shipment_df[shipment_df["Shipment Job"].isin(kept)]
supp_f = build_supplier_summary(fdata, fship) if not fdata.empty else supp_sum

# ── KPIs ──────────────────────────────────────────────────────────────────────

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Shipments", f"{bill_sum_f['Shipment Job'].nunique():,}")
k2.metric("Containers", f"{int(ship_sum_f['Container Count'].fillna(0).sum()):,}"
          if "Container Count" in ship_sum_f.columns else "—")
k3.metric(f"{BASE_CCY} charges",
          f"{bill_det_f[bill_det_f['Currency'] == BASE_CCY]['Total'].sum():,.2f}")
k4.metric(f"{FX_CCY} charges",
          f"{bill_det_f[bill_det_f['Currency'] == FX_CCY]['Total'].sum():,.2f}")
k5.metric(f"Total ({BASE_CCY})", f"{bill_det_f['Local Total'].sum():,.2f}")

# ── Tabs ──────────────────────────────────────────────────────────────────────

tab1, tab2, tab3, tab4, tab6, tab5 = st.tabs([
    "Shipment summary", "Billing summary", "Billing detail", "Analysis",
    "Charge by container", "Download",
])

with tab1:
    st.markdown('<div class="section-title">Shipment summary</div>', unsafe_allow_html=True)
    st.dataframe(ship_sum_f, use_container_width=True, hide_index=True)
    containers = (int(ship_sum_f["Container Count"].fillna(0).sum())
                  if "Container Count" in ship_sum_f.columns else 0)
    st.caption(f"{len(ship_sum_f)} shipments · {containers} containers")

with tab2:
    st.markdown('<div class="section-title">Billing summary — per job</div>', unsafe_allow_html=True)
    st.dataframe(bill_sum_f, use_container_width=True, hide_index=True)
    base_total = bill_sum_f[BASE_CCY].sum() if BASE_CCY in bill_sum_f.columns else 0
    fx_total = bill_sum_f[FX_CCY].sum() if FX_CCY in bill_sum_f.columns else 0
    local_total = bill_sum_f[LOCAL_LABEL].sum() if LOCAL_LABEL in bill_sum_f.columns else 0
    st.caption(f"{BASE_CCY} {base_total:,.2f}  ·  {FX_CCY} {fx_total:,.2f}  "
               f"·  {LOCAL_LABEL} {local_total:,.2f}")

with tab3:
    st.markdown('<div class="section-title">Billing detail — all charge lines</div>', unsafe_allow_html=True)
    st.dataframe(bill_det_f, use_container_width=True, hide_index=True)
    st.caption(f"{len(bill_det_f):,} charge lines")

with tab4:
    if bill_sum_f.empty:
        st.warning("No shipments match the current filters. Clear a filter to see data.")
    else:
        st.markdown(f'<div class="section-title">Total billed by supplier ({BASE_CCY})</div>',
                    unsafe_allow_html=True)
        fig1 = px.bar(
            supp_f.sort_values(SUPPLIER_TOTAL_COL),
            x=SUPPLIER_TOTAL_COL, y="Supplier Name", orientation="h",
            color_discrete_sequence=[PRIMARY],
            labels={SUPPLIER_TOTAL_COL: LOCAL_LABEL, "Supplier Name": ""},
            text=SUPPLIER_TOTAL_COL,
        )
        fig1.update_traces(texttemplate="%{text:,.0f}", textposition="outside",
                           marker_line_width=0, textfont_color=TEXT, marker_color=PRIMARY)
        fig1.update_layout(**PLOTLY_LAYOUT, xaxis_title="", yaxis_title="")
        st.plotly_chart(fig1, use_container_width=True)

        r2l, r2r = st.columns(2)

        with r2l:
            st.markdown(f'<div class="section-title">Top 10 charge types ({BASE_CCY})</div>',
                        unsafe_allow_html=True)
            top10 = (bill_det_f.groupby("Description")["Local Total"].sum()
                     .sort_values(ascending=False).head(10).index.tolist())
            charge_stack = (bill_det_f[bill_det_f["Description"].isin(top10)]
                            .groupby(["Description", "Currency"])["Local Total"].sum().reset_index())
            charge_stack["rank"] = charge_stack["Description"].map(
                {d: i for i, d in enumerate(reversed(top10))})
            charge_stack = charge_stack.sort_values("rank")
            fig3 = px.bar(
                charge_stack, x="Local Total", y="Description", color="Currency",
                orientation="h", color_discrete_map={BASE_CCY: ACCENT, FX_CCY: PRIMARY},
                barmode="stack", labels={"Local Total": LOCAL_LABEL, "Description": ""},
            )
            fig3.update_layout(**PLOTLY_LAYOUT, xaxis_title="", yaxis_title="")
            fig3.update_traces(marker_line_width=0)
            st.plotly_chart(fig3, use_container_width=True)
            st.caption(f"Non-{BASE_CCY} charges are shown converted to {BASE_CCY}.")

        with r2r:
            st.markdown('<div class="section-title">Shipments by Incoterms</div>', unsafe_allow_html=True)
            if "Incoterms" in ship_sum_f.columns:
                inco_data = ship_sum_f.groupby("Incoterms")["Shipment Job"].nunique().reset_index()
                inco_data.columns = ["Incoterms", "Shipments"]
                fig4 = px.pie(inco_data, values="Shipments", names="Incoterms",
                              color_discrete_sequence=PALETTE, hole=0.45)
                fig4.update_layout(**PLOTLY_LAYOUT)
                fig4.update_traces(textfont_color="#ffffff", textfont_size=13)
                st.plotly_chart(fig4, use_container_width=True)

        st.markdown(f'<div class="section-title">Monthly spend by Incoterm ({BASE_CCY})</div>',
                    unsafe_allow_html=True)
        bd_m2, month_order2 = add_month_col(bill_det_f, ship_sum_f)
        if month_order2 and "Incoterms" in bd_m2.columns:
            inco_monthly = bd_m2.groupby(["Month", "Incoterms"])["Local Total"].sum().reset_index()
            inco_monthly = inco_monthly[inco_monthly["Local Total"] > 0]
            inco_monthly["Month"] = pd.Categorical(
                inco_monthly["Month"], categories=month_order2, ordered=True)
            inco_monthly = inco_monthly.sort_values("Month")
            fig5 = px.line(
                inco_monthly, x="Month", y="Local Total", color="Incoterms",
                color_discrete_sequence=PALETTE, markers=True,
                labels={"Local Total": LOCAL_LABEL, "Month": ""},
            )
            fig5.update_layout(**PLOTLY_LAYOUT, xaxis_title="", yaxis_title=LOCAL_LABEL)
            fig5.update_traces(line_width=2.5, marker_size=8)
            st.plotly_chart(fig5, use_container_width=True)
            st.caption("Each line tracks total spend per Incoterm by ETD month.")
        else:
            st.info("Add shipments with ETD dates and Incoterms to see the monthly trend.")

        st.markdown('<div class="section-title">Supplier summary</div>', unsafe_allow_html=True)
        st.dataframe(supp_f, use_container_width=True, hide_index=True)

with tab6:
    st.markdown('<div class="section-title">Charge by container</div>', unsafe_allow_html=True)
    st.caption(
        "Every charge line on the selected shipments, allocated evenly across that "
        f"job's containers. Amounts are GST-exclusive, in {BASE_CCY}, and each "
        "charge's shares add back to the invoiced figure to the cent."
    )

    invoice_source = fdata if not fdata.empty else data
    cbc, diag = build_charge_by_container(invoice_source, shipment_df)

    if diag["missing_columns"]:
        st.error("The billing export is missing: " + ", ".join(diag["missing_columns"]))
    elif cbc.empty:
        st.warning("No charge lines to allocate for the current filters.")
    else:
        if diag["invoice_number_source"]:
            st.caption(f"Invoice number read from the '{diag['invoice_number_source']}' column.")
        else:
            st.warning(
                "No invoice number column found in the billing export, so rows are "
                "grouped by job only and the invoice columns are blank. Add the invoice "
                "number to the CargoWise export to populate them."
            )
        if diag["jobs_without_containers"]:
            st.info(
                f"{len(diag['jobs_without_containers'])} job(s) have no container listed "
                "in the shipment report; their charges sit on a single "
                f"'{NO_CONTAINER_LABEL}' row so nothing is dropped."
            )

        template = st.file_uploader(
            "Optional: upload an existing report to copy its column order",
            type=["csv"], key="cbc_template")
        if template is not None:
            try:
                template_cols = list(pd.read_csv(template, nrows=0).columns)
                cbc = reorder_to_template(cbc, template_cols)
            except Exception as e:
                st.error(f"Could not read the template header: {e}")

        code_count = len([c for c in cbc.columns if c not in (
            JOB_LABEL, CONTAINER_LABEL, COMMERCIAL_REF_LABEL,
            INVOICE_NUMBER_LABEL, INVOICE_DATE_LABEL, "TOTAL")])
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Rows", f"{len(cbc):,}")
        m2.metric("Containers", f"{cbc[CONTAINER_LABEL].nunique():,}")
        m3.metric("Charge codes", f"{code_count:,}")
        m4.metric(f"Allocated ({BASE_CCY})", f"{cbc['TOTAL'].sum():,.2f}")

        st.dataframe(cbc, use_container_width=True, hide_index=True)

        st.download_button(
            label="Download charge by container (CSV)",
            data=cbc.to_csv(index=False).encode("utf-8"),
            file_name=INVOICE_CSV_FILENAME,
            mime="text/csv",
        )

        # Reconciliation: allocated total must equal the source charge total.
        source_total = pd.to_numeric(
            invoice_source[CHARGE_AMOUNT_COL], errors="coerce").fillna(0).sum()
        difference = round(cbc["TOTAL"].sum() - source_total, 2)
        if abs(difference) < 0.01:
            st.success(f"Reconciled: allocated total matches the billing export "
                       f"({source_total:,.2f} {BASE_CCY}).")
        else:
            st.error(f"Allocated total is out by {difference:,.2f} {BASE_CCY} versus the "
                     f"billing export ({source_total:,.2f}). Check for charge lines "
                     "whose job is missing from the shipment report.")

with tab5:
    st.markdown('<div class="section-title">Download consolidated report</div>', unsafe_allow_html=True)
    active = any([f_supplier, f_origin, f_dest, f_inco, f_mode, f_month])
    scope = (f"**{len(bill_sum_f)} of {len(bill_sum)} shipments** match the current filters."
             if active else f"**All {len(bill_sum)} billed shipments** are included.")
    st.info(scope)
    st.download_button(
        label="Download Excel report",
        data=create_report(ship_sum_f, bill_sum_f, bill_det_f, supp_f),
        file_name=REPORT_FILENAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown(f"""
**Sheets included:**
- **Shipment Summary** — one row per shipment
- **Billing Summary** — per-job {BASE_CCY} / {FX_CCY} / {LOCAL_LABEL}
- **Billing Detail** — every charge line
- **Supplier Summary** — shipment count and spend per supplier
- **Analysis** — key metrics and spend breakdowns
    """)

with st.sidebar:
    st.download_button(
        label="Download report",
        data=create_report(ship_sum_f, bill_sum_f, bill_det_f, supp_f),
        file_name=REPORT_FILENAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(f"{len(bill_sum_f)} shipments in the current export")
